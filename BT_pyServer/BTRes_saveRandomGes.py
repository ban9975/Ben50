from pickle import NONE
import interface

from openpyxl import *
from datetime import datetime

interf = interface.interface()

def main():
    fileName = 'wristband/v4/adi_res_test.xlsx'
    gestures = ['down', 'up', 'thumb', 'little finger', 'stretch', 'fist', 'rest']
    workbook = load_workbook(fileName)
    start = datetime.now()
    mode = int(input("0: calibration, 1: random: "))
    stretch = [285.33, 275.42, 336.76, 244.03]
    gnd = [355, 377, 363, 386]
    off = [s-g for s,g in zip(stretch,gnd)]
    print(off)
    if mode == 1:
        worksheet = workbook.create_sheet("random")
        title = ['gesture', 'start', 'end', 0, 1, 2, 3]
        for i in range(len(title)):
            worksheet.cell(row=1, column=i+1, value=title[i])
        for i in range(4):
            worksheet.cell(row=2, column=i+4, value=off[i])
    elif mode == 0:
        stretch = [0, 0, 0, 0]
        off = [0, 0, 0, 0]
    row = 3
    while True:
        gesture = (row-3)%7
        _input = input("{}\tinput: ".format(gestures[gesture]))
        if _input == 'b':
            row -= 1
        elif _input == 'e':
            interf.end_process()
            if mode == 1:
                workbook.save(fileName)
                workbook.close()
            elif mode == 0:
                for k in range(4):
                    stretch[k] = round(stretch[k]/(row - 3),2)
                print(stretch)
            break
        else:
            interf.write(str(1))
            if mode == 1:
                worksheet.cell(row=row, column=1, value=gesture)
                worksheet.cell(row=row, column=2, value=str(datetime.now()-start))

            avg = [0, 0, 0, 0]
            for j in range(20):
                for k in range(4):
                    btIn = float(interf.read())
                    while btIn == 0:
                        print(0)
                        btIn = float(interf.read())
                    avg[k] += btIn
            print(row-2, end='\t')
            for k in range(4):
                avg[k] /= 20
                res = 300 * avg[k] / (5000 - avg[k] * 3) - off[k]
                print(round(res, 2), end='\t')
                if mode == 1:
                    worksheet.cell(row=row, column=k+4, value=round(res, 2))
                    workbook.save(fileName)
                elif mode == 0:
                    stretch[k] += res

            print()
            if mode == 1:
                worksheet.cell(row=row, column=3, value=str(datetime.now()-start))
            row += 1

if __name__ == '__main__':
    main()