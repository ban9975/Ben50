from pickle import NONE
import interface

# import numpy as np
# import time
# import sys
# import os

from openpyxl import *
from datetime import datetime

interf = interface.interface()

def main():
    fileName = 'wristband/small_adi.xlsx'
    gestures = ['down', 'up', 'thumb', 'little finger', 'stretch', 'fist', 'rest', 'random']
    workbook = load_workbook(fileName)
    start = datetime.now()
    gesture = int(input("gesture: "))
    name = "{}_".format(gestures[gesture])
    print(name)
    worksheet = workbook.create_sheet(name)
    title = ['gesture', 'start', 'end', 0, 1, 2, 3]
    for i in range(len(title)):
        worksheet.cell(row=1, column=i+1, value=title[i])
    row = 2
    stretch = [425, 367, 400, 421]
    gnd = [355, 377, 363, 386]
    off = [s-g for s,g in zip(stretch,gnd)]
    print(off)
    while True:
        _input = input("input: ")
        if _input == 'b':
            row -= 1
        elif _input == 'e':
            interf.end_process()
            workbook.save(fileName)
            workbook.close()
            break
        else:
            if _input == '':
                _input = gesture
            interf.write(str(1))
            worksheet.cell(row=row, column=1, value=int(_input))
            worksheet.cell(row=row, column=2, value=str(datetime.now()-start))

            avg = [0, 0, 0, 0]
            for j in range(20):
                for k in range(4):
                    btIn = float(interf.read())
                    while btIn == 0:
                        print(0)
                        btIn = float(interf.read())
                    # if k == 0:
                    #     print(btIn * 3)
                    # ADC conversion
                    avg[k] += btIn
            print(row-1, end='\t')
            for k in range(4):
                avg[k] /= 20
                res = 300 * avg[k] / (5000 - avg[k] * 3) - off[k]
                print(round(res, 2), end='\t')
                worksheet.cell(row=row, column=k+4, value=round(res, 2))
            print()
            worksheet.cell(row=row, column=3, value=str(datetime.now()-start))
            row += 1

if __name__ == '__main__':
    main()