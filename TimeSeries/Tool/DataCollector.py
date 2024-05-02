from Constants import *
from BTController import *
import os
from openpyxl import *
import pandas as pd
import matplotlib.pyplot as plt


class DataCollector:
    def __init__(self) -> None:
        self.fileName = ""
        self.bt = BTController()

    def connectBT(self, port: str) -> None:
        self.bt.do_connect(port)

    def prepareFile(self) -> None:
        if not os.path.exists(self.fileName):
            workbook = Workbook()
            workbook.save(self.fileName)
            workbook.close()
        self.workbook = load_workbook(self.fileName)

    def startCollect(self, fileName: str, sequence: str, iteration: int):
        self.fileName = os.path.expanduser(f"{fileName}.xlsx")
        self.prepareFile()
        self.worksheet = self.workbook.create_sheet(f"{sequence}_{iteration}")
        self.sheetName = self.workbook.sheetnames[-1]
        title = ["gesture", "0", "1", "2"]
        for i in range(nSensors + 1):
            self.worksheet.cell(row=1, column=i + 1, value=title[i])
        self.row = 2
        timeLen = str(len(sequence) * iteration * 4 + 2)
        self.bt.write(timeLen)
        self.currentValue = float(self.bt.read())
        return self.currentValue == 2048

    def collectData(self, ges: int) -> bool:
        if self.currentValue == 2048:
            return True
        data = [self.currentValue]
        for k in range(nSensors - 1):
            btIn = float(self.bt.read())
            while btIn == 0:
                print(0)
                btIn = float(self.bt.read())
            data.append(btIn)
        data = [
            round(3000 * data[i] / (5000 - data[i] * 3), 2) for i in range(nSensors)
        ]
        self.worksheet.cell(row=self.row, column=1, value=ges)
        for i in range(nSensors):
            self.worksheet.cell(row=self.row, column=2 + i, value=data[i])
        self.row += 1
        self.currentValue = float(self.bt.read())
        return False

    def stopCollect(self):
        if "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])
        self.workbook.save(self.fileName)
        self.workbook.close()

    def plotData(self, plotPath: str) -> str:
        xls = pd.ExcelFile(self.fileName)
        data = xls.parse(self.sheetName)
        plt.figure()
        plt.ylim(500, 3200)
        plt.title(f"{self.fileName} {self.sheetName}")
        for col in data.columns:
            if col == "gesture":
                plt.plot(
                    [i * 10 for i in range(len(data[col]))],
                    [i * 50 + 3000 for i in data[col]],
                    label="gesture",
                )
                continue
            plt.plot(
                [i * 10 for i in range(len(data[col]))],
                data[col],
                label=f"Sensor {col}",
            )
        plt.ylabel("resistance(ohm)")
        plt.xlabel("time(ms)")
        plt.legend(loc="upper right")
        figFolder = os.path.expanduser(
            os.path.join(
                plotPath,
                os.path.basename(os.path.normpath(self.fileName)).split(".")[0],
            )
        )
        figName = os.path.expanduser(
            os.path.join(
                figFolder,
                f"{self.sheetName}.png",
            )
        )
        print(figFolder)
        if not os.path.exists(figFolder):
            os.makedirs(figFolder, exist_ok=True)
        plt.savefig(figName)
        return figName
