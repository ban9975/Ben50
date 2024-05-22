import sys
import os
import random

parent_dir = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
sys.path.append(parent_dir)
from BT_common.BTController import BTController
import time
from openpyxl import *
import os

gestures = ["down", "up", "open", "little"]
title = ["gesture", "0", "1", "2", "3"]
nSensor = 3
bt = BTController()
bt.do_connect("COM23")
# bt.do_connect("/dev/cu.H-C-2010-06-01")
fileName = os.path.join(
    os.getcwd(), "Excel_data/v8/Time_series", f'{input("File name: ")}.xlsx'
)
if not os.path.exists(fileName):
    workbook = Workbook()
    workbook.save(fileName)
    workbook.close()
workbook = load_workbook(fileName)
while True:
    seq = input("sequence=?")
    if seq == "e":
        break
    iter = int(input("iterate=?"))
    worksheet = workbook.create_sheet(f"{seq}_{iter}")
    for i in range(len(title)):
        worksheet.cell(row=1, column=i + 1, value=title[i])
    row = 2
    # print("ready")
    # time.sleep(1)
    gestureList = []
    for i in range(iter):
        for s in seq:
            gestureList.append(int(s))
    t = str(len(gestureList) * 4 + 2)
    bt.write(t)
    section = 1
    ges = -1  # neutral
    j = 0
    print("neutral")
    val = float(bt.read())
    start = time.time()
    while val != 2048:
        # print(time.time(), start)
        if time.time() - start > 2:
            if section % 2 == 1 and section < len(gestureList) * 2:
                ges = gestureList[j]
                j += 1
                print(gestures[ges])
            elif section % 2 == 0 or section >= len(gestureList) * 2:
                ges = -1
                print("neutral")
            section += 1
            start = time.time()
        data = [val]
        for k in range(nSensor - 1):
            btIn = float(bt.read())
            while btIn == 0:
                print(0)
                btIn = float(bt.read())
            data.append(btIn)
        data = [round(3000 * data[i] / (5000 - data[i] * 3), 2) for i in range(nSensor)]
        worksheet.cell(row=row, column=1, value=ges)
        for i in range(nSensor):
            worksheet.cell(row=row, column=2 + i, value=data[i])
        row += 1
        val = float(bt.read())
    print("stop")
    if "Sheet" in workbook.sheetnames:
        workbook.remove(workbook["Sheet"])
    workbook.save(fileName)
