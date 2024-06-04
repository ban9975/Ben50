import sys
import os
import random

parent_dir = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
sys.path.append(parent_dir)
import time
from openpyxl import *
import os
import serial


class BTController:
    def __init__(self) -> None:
        pass

    def do_connect(self, port):
        while not self.connect(port):
            print("connecting")
            self.connect(port)

    def connect(self, port):
        try:
            self.ser = serial.Serial(port, 9600, timeout=2)
            print("connect success")
            print(port)
            self.write("0")
            return True
        except serial.serialutil.SerialException as ex:
            print("fail to connect")
            print(ex)
            print("")
            return False

    def __del__(self):
        if self.ser.is_open:
            self.ser.close()

    def write(self, output):
        send = output.encode("utf-8")
        self.ser.write(send)

    def read(self):
        waiting = self.ser.in_waiting
        if waiting >= 0:
            rv = self.ser.read(2)
            rv = int.from_bytes(rv, byteorder="big", signed=False)
            return rv
    
    def reset(self):
        self.ser.reset_input_buffer()


gestures = ["down", "up", "open", "little"]
title = ["gesture", "0", "1", "2", "3"]
nSensor = 3
bt = BTController()
bt.do_connect("COM19")
# bt.do_connect("/dev/cu.H-C-2010-06-01")
# bt.do_connect("/dev/cu.hc05")
fileName = os.path.join(os.getcwd(), f'{input("File name: ")}.xlsx')
if not os.path.exists(fileName):
    workbook = Workbook()
    workbook.save(fileName)
    workbook.close()
workbook = load_workbook(fileName)
while True:
    seq = input("sequence=?")
    if seq == "e":
        break
    iter = int(input("iterate=?"))
    worksheet = workbook.create_sheet(f"{seq}_{iter}")
    for i in range(len(title)):
        worksheet.cell(row=1, column=i + 1, value=title[i])
    row = 2
    # print("ready")
    # time.sleep(1)
    gestureList = []
    for i in range(iter):
        for s in seq:
            gestureList.append(int(s))
    t = str(len(gestureList) * 4 + 2)
    bt.reset()
    bt.write(t)
    section = 1
    ges = -1  # neutral
    j = 0
    print("neutral")
    val = float(bt.read())
    print(val)
    start = time.time()
    while val != 2048:
        # print(time.time(), start)
        if time.time() - start > 2:
            if section % 2 == 1 and section < len(gestureList) * 2:
                ges = gestureList[j]
                j += 1
                print(gestures[ges])
            elif section % 2 == 0 or section >= len(gestureList) * 2:
                ges = -1
                print("neutral")
            section += 1
            start = time.time()
        data = [val]
        for k in range(nSensor - 1):
            btIn = float(bt.read())
            while btIn == 0:
                print(0)
                btIn = float(bt.read())
            data.append(btIn)
        data = [round(3000 * data[i] / (5000 - data[i] * 3), 2) for i in range(nSensor)]
        worksheet.cell(row=row, column=1, value=ges)
        for i in range(nSensor):
            worksheet.cell(row=row, column=2 + i, value=data[i])
        row += 1
        val = float(bt.read())
    print("stop")
    if "Sheet" in workbook.sheetnames:
        workbook.remove(workbook["Sheet"])
    workbook.save(fileName)
