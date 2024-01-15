import sys
import os
import random

parent_dir = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
sys.path.append(parent_dir)
from BT_common.BTController import BTController
import time
from openpyxl import *
import os

gestures = ["down", "up", "open", "close"]
title = ["gesture", "0", "1", "2"]
nSensor = 3
bt = BTController()
bt.do_connect("COM12")
fileName = os.path.join(
    os.getcwd(), "Excel_data/v8/Time_series", f'{input("File name: ")}.xlsx'
)
if not os.path.exists(fileName):
    workbook = Workbook()
    workbook.save(fileName)
    workbook.close()
workbook = load_workbook(fileName)
while True:
    t = input("start?")
    if t == "e":
        break    
    t = "28"
    worksheet = workbook.create_sheet()
    for i in range(len(title)):
        worksheet.cell(row=1, column=i + 1, value=title[i])
    row = 2
    print("ready")
    time.sleep(1)
    bt.write(t)
    start = time.time()
    section = 0
    ges = 999  # neutral
    gestureList = [0, 1, 2,0,1,2]
    random.shuffle(gestureList)
    j = 0
    val = float(bt.read())
    while val != 2048:
        # print(time.time(), start)
        if time.time() - start > 2:
            if section % 2 == 1 and section < len(gestureList) * 2:
                ges = gestureList[j]
                j += 1
                print(gestures[ges])
            elif section % 2 == 0 or section >= len(gestureList) * 2:
                ges = 999
                print("neutral")
            section += 1
            start = time.time()
        data = [val]
        for k in range(nSensor - 1):
            btIn = float(bt.read())
            while btIn == 0:
                print(0)
                btIn = float(bt.read())
            if btIn == 2048:
                print("missed")
            data.append(btIn)
        data = [round(3000 * data[i] / (5000 - data[i] * 3), 2) for i in range(nSensor)]
        worksheet.cell(row=row, column=1, value=ges)
        worksheet.cell(row=row, column=2, value=data[0])
        worksheet.cell(row=row, column=3, value=data[1])
        worksheet.cell(row=row, column=4, value=data[2])
        row += 1
        val = float(bt.read())
    print("stop")
    workbook.save(fileName)
