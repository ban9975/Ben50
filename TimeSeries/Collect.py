import sys
import os
parent_dir = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
print(parent_dir)
sys.path.append(parent_dir)
from BT_common.BTController import BTController
import time
from openpyxl import *
import os

gestures = ["down", "up", "open", "close"]
title = ["0", "1", "2"]
nSensor = 3
bt = BTController()
bt.do_connect("COM12")
fileName = os.path.join(
    os.getcwd(), "Excel_data/v8/Time_series", f'{input("File name: ")}.xlsx'
)
if not os.path.exists(fileName):
    workbook = Workbook()
    workbook.save(fileName)
    workbook.close()
workbook = load_workbook(fileName)
while True:
    gesture = input("Gesture:")
    if gesture == "e":
        break
    else:
        gesture = int(gesture)
    worksheet = workbook.create_sheet(gestures[gesture])
    for i in range(len(title)):
        worksheet.cell(row=1, column=i + 1, value=title[i])
    t = [2, 2, 2]
    words = ["neutral", gestures[gesture], "neutral"]
    row = 2
    print("ready")
    time.sleep(1)
    print(words[0])
    bt.write(str(sum(t)))
    val = float(bt.read())
    section = 0
    start = time.time()
    while val != 2048:
        if section<len(t)-1 and time.time() - start > t[section]:
            section += 1
            print(words[section])
            start = time.time()
        data = [val]
        for k in range(nSensor - 1):
            btIn = float(bt.read())
            while btIn == 0:
                print(0)
                btIn = float(bt.read())
            data.append(btIn)
        data = [round(3000 * data[i] / (5000 - data[i] * 3), 2) for i in range(nSensor)]
        worksheet.cell(row=row, column=1, value=data[0])
        worksheet.cell(row=row, column=2, value=data[1])
        worksheet.cell(row=row, column=3, value=data[2])
        row += 1
        val = float(bt.read())
    print("stop")
    workbook.save(fileName)