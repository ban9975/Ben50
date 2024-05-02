import sys
import os
from openpyxl import *
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
import time
import cv2
from datetime import datetime
from BTController import bt, list_all_ports

from Rf import modes, caliCnt, Rf, gestures, nSensor
from Plot import Plot

root = os.getcwd()
class Tool(QMainWindow):
    def __init__(self):
        super().__init__()
        self.widget = QWidget()
        self.layout = QVBoxLayout(self.widget)
        self.btWidget = QWidget()
        self.btLayout = QHBoxLayout(self.btWidget)
        self.btLabel = QLabel("COM Port:")
        self.btCombo = QComboBox()
        self.btCombo.addItems(list_all_ports())
        self.btBtn = QPushButton("Connect")
        self.btBtn.clicked.connect(self.connectBT)
        self.btLayout.addWidget(self.btLabel)
        self.btLayout.addWidget(self.btCombo)
        self.btLayout.addWidget(self.btBtn)
        self.layout.addWidget(self.btWidget)
        self.tab = QTabWidget()
        self.tab.addTab(TrainPage(), "Train RF")
        self.tab.addTab(PredictPage(), "Predict")
        self.tab.addTab(CollectPage(), "Collect")
        self.layout.addWidget(self.tab)
        self.setCentralWidget(self.widget)
        self.setFixedSize(1800, 900)
        self.show()

    def connectBT(self):
        bt.do_connect(self.btCombo.currentText())


class TrainPage(QWidget):
    def __init__(self):
        super().__init__()
        self.trainLabel = QLabel("Train data path:")
        self.trainPath = QLabel(" ")
        self.trainBtn = QPushButton("Select train data")
        self.trainBtn.clicked.connect(lambda: self.selectFile(self.trainPath))
        self.train = QWidget()
        self.trainLayout = QHBoxLayout(self.train)
        self.trainLayout.addWidget(self.trainLabel)
        self.trainLayout.addWidget(self.trainPath)
        self.trainLayout.addWidget(self.trainBtn)
        self.testLabel = QLabel("Test data path:")
        self.testPath = QLabel("")
        self.testBtn = QPushButton("Select test data")
        self.testBtn.clicked.connect(lambda: self.selectFile(self.testPath))
        self.test = QWidget()
        self.testLayout = QHBoxLayout(self.test)
        self.testLayout.addWidget(self.testLabel)
        self.testLayout.addWidget(self.testPath)
        self.testLayout.addWidget(self.testBtn)
        self.mode = QComboBox()
        self.mode.addItems(modes)
        self.runBtn = QPushButton("Run")
        self.runBtn.clicked.connect(self.run)
        self.result = QLabel()
        self.saveBtn = QPushButton("Save model")
        self.saveBtn.clicked.connect(self.saveModel)
        self.plotTitleLabel = QLabel('Plot title: ')
        self.plotTitleEdit = QLineEdit()
        self.plotDropdown = QComboBox()
        self.plotDropdown.addItems(['3D', 'Sensor', 'Confusion matrix'])
        self.plotBtn = QPushButton("Plot")
        self.plotBtn.clicked.connect(self.plot)
        self.layout = QGridLayout(self)
        self.layout.addWidget(self.train, 0, 0, 1, 4)
        self.layout.addWidget(self.test, 1, 0, 1, 4)
        self.layout.addWidget(self.mode, 2, 1, 1, 1)
        self.layout.addWidget(self.runBtn, 2, 2, 1, 1)
        self.layout.addWidget(self.result, 3, 0, 2, 4)
        self.layout.addWidget(self.saveBtn, 4, 2, 1, 1)
        # self.layout.addWidget(self.plotTitleLabel, 5, 0, 1, 1)
        # self.layout.addWidget(self.plotTitleEdit, 5, 1, 1, 1)
        self.layout.addWidget(self.plotDropdown, 5, 2, 1, 1)
        self.layout.addWidget(self.plotBtn, 5, 3, 1, 1)

    def selectFile(self, label: QLabel):
        dialog = QFileDialog(self)
        dialog.setDirectory(os.path.join(root,'Excel_data'))
        dialog.setFileMode(QFileDialog.FileMode.ExistingFiles)
        dialog.setViewMode(QFileDialog.ViewMode.List)
        if dialog.exec():
            fileName = dialog.selectedFiles()[0]
            label.setText(fileName)

    def run(self):
        self.rf = Rf()
        self.rf.prepareData(
            self.trainPath.text(), self.testPath.text(), self.mode.currentIndex()
        )
        self.result.setText(self.rf.runRf())

    def saveModel(self):
        fileName = QFileDialog.getSaveFileName(directory=os.path.join(root,'Model'))
        self.rf.saveModel(fileName[0])

    def plot(self):
        self.plot = Plot()
        fileName = QFileDialog.getSaveFileName(directory=os.path.join(root,'Wristband_plots/versions'))
        if self.plotDropdown.currentText() == '3D':
            self.plot.plot_3d(self.trainPath.text(), self.testPath.text(), self.mode.currentIndex(), fileName[0])
        elif self.plotDropdown.currentText() == 'Sensor':
            self.plot.plot_sensor(self.trainPath.text(), self.testPath.text(), self.mode.currentIndex(), fileName[0])
        elif self.plotDropdown.currentText() == 'Confusion matrix':
            self.rf.confusionMatrix(fileName[0])

class PredictPage(QWidget):
    def __init__(self):
        super().__init__()
        self.modelLabel = QLabel("Model path:")
        self.modelPath = QLabel(" ")
        self.modelBtn = QPushButton("Select model")
        self.modelBtn.clicked.connect(lambda: self.selectFile(self.modelPath))
        self.model = QWidget()
        self.modelLayout = QHBoxLayout(self.model)
        self.modelLayout.addWidget(self.modelLabel)
        self.modelLayout.addWidget(self.modelPath)
        self.modelLayout.addWidget(self.modelBtn)
        self.mode = QComboBox()
        self.mode.addItems(modes)
        self.runBtn = QPushButton("Run")
        self.runBtn.clicked.connect(self.run)
        self.nextBtn = QPushButton("Calibrate next")
        self.nextBtn.setDisabled(True)
        self.stopBtn = QPushButton("Stop")
        self.stopBtn.setDisabled(True)
        self.result = QLabel()
        self.values = QLabel()
        self.resultFig = QLabel()
        self.camera = QLabel()
        self.layout = QGridLayout(self)
        self.layout.addWidget(self.model, 0, 0, 1, 3)
        self.layout.addWidget(self.mode, 1, 1, 1, 1)
        self.layout.addWidget(self.runBtn, 1, 2, 1, 1)
        self.layout.addWidget(self.nextBtn, 2, 1, 1, 1)
        self.layout.addWidget(self.stopBtn, 2, 2, 1, 1)
        self.layout.addWidget(self.result, 3, 0, 1, 1)
        self.layout.addWidget(self.values, 3, 1, 1, 2)
        self.layout.addWidget(self.resultFig, 4, 0, 1, 1)
        self.layout.addWidget(self.camera, 4, 2, 1, 1)
        self.calikey = QShortcut(QKeySequence("Return"), self)
        self.calikey.activated.connect(self.caliNext)
        self.cap = cv2.VideoCapture(0)
        self.camTimer = QTimer()
        self.camTimer.timeout.connect(self.show_camera)
        self.camTimer.start(30)

    def selectFile(self, label: QLabel):
        dialog = QFileDialog(self, directory=os.path.join(root,'Model'))
        dialog.setFileMode(QFileDialog.FileMode.ExistingFiles)
        dialog.setViewMode(QFileDialog.ViewMode.List)
        if dialog.exec():
            fileName = dialog.selectedFiles()[0]
            label.setText(fileName)

    def run(self):
        self.disableBtns(False)
        self.thread = QThread()
        self.worker = PredictWorker(self.modelPath.text())
        self.nextBtn.clicked.connect(lambda:self.worker.nextSignal.emit())
        self.stopBtn.clicked.connect(lambda:self.worker.stopSignal.emit())
        self.worker.resultSignal.connect(
            lambda result, values: self.setLabel(result, values)
        )
        self.worker.predictSignal.connect(lambda: self.nextBtn.setDisabled(True))
        self.worker.finishSignal.connect(lambda: self.disableBtns(True))
        self.worker.finishSignal.connect(self.worker.deleteLater)
        self.worker.finishSignal.connect(self.thread.quit)
        self.thread.finished.connect(self.thread.deleteLater)
        self.worker.moveToThread(self.thread)
        self.thread.started.connect(self.worker.work)
        self.thread.start()

    def setLabel(self, result, values):
        self.result.setText(result)
        self.resultFig.setPixmap(
            QPixmap(os.path.join(os.getcwd(), "Tool", "gesture_figs", f"{result}.png"))
        )
        self.values.setText(values)

    def caliNext(self):
        if self.nextBtn.isEnabled():
            self.worker.nextSignal.emit()

    def disableBtns(self, state):
        self.nextBtn.setDisabled(state)
        self.stopBtn.setDisabled(state)

    def show_camera(self):
        _, self.image = self.cap.read()
        show = cv2.resize(self.image, (480, 320))
        show = cv2.cvtColor(show, cv2.COLOR_BGR2RGB)
        showImage = QImage(
            show.data, show.shape[1], show.shape[0], QImage.Format_RGB888
        )
        self.camera.setPixmap(QPixmap.fromImage(showImage))


class PredictWorker(QObject):
    resultSignal = pyqtSignal((str, str))
    nextSignal = pyqtSignal()
    predictSignal = pyqtSignal()
    stopSignal = pyqtSignal()
    finishSignal = pyqtSignal()

    def __init__(self, modelPath):
        super().__init__()
        self.rf = Rf()
        self.next = False
        self.modelPath = modelPath
        self.stop = False
        self.nextSignal.connect(lambda: setattr(self, "next", True))
        self.stopSignal.connect(lambda: setattr(self, "stop", True))

    def work(self):
        self.rf.predictSetup(self.modelPath)
        for i in range(caliCnt):
            while not self.next:
                pass
            self.resultSignal.emit(f"Calibration {i+1}", self.rf.calibration())
            self.next = False
        self.predictSignal.emit()
        while not self.stop:
            self.resultSignal.emit(*self.rf.predict())
            time.sleep(0.5)
        self.finishSignal.emit()


class CollectPage(QWidget):
    def __init__(self):
        super().__init__()
        self.filePath = QLabel("")
        self.fileBtn = QPushButton("Select store path")
        self.fileBtn.clicked.connect(self.selectFile)
        self.gestureCombo = QComboBox()
        self.gestureCombo.addItems(gestures)
        self.collectBtn = QPushButton("Collect")
        self.collectBtn.clicked.connect(
            lambda: self.collect(self.gestureCombo.currentIndex())
        )
        self.nextBtn = QPushButton("Next")
        self.nextBtn.setDisabled(True)
        self.backBtn = QPushButton("Back")
        self.backBtn.setDisabled(True)
        self.stopBtn = QPushButton("Stop")
        self.stopBtn.setDisabled(True)
        self.gesture = QLabel()
        self.values = QLabel()
        self.layout = QGridLayout(self)
        self.layout.addWidget(self.filePath, 0, 0, 1, 2)
        self.layout.addWidget(self.fileBtn, 0, 2, 1, 1)
        self.layout.addWidget(self.gestureCombo, 1, 0, 1, 1)
        self.layout.addWidget(self.collectBtn, 1, 1, 1, 1)
        self.layout.addWidget(self.nextBtn, 2, 0, 1, 1)
        self.layout.addWidget(self.backBtn, 2, 1, 1, 1)
        self.layout.addWidget(self.stopBtn, 2, 2, 1, 1)
        self.layout.addWidget(self.gesture, 3, 0, 1, 1)
        self.layout.addWidget(self.values, 3, 1, 1, 2)
        self.nextkey = QShortcut(QKeySequence("Return"), self)
        self.nextkey.activated.connect(self.next)

    def selectFile(self):
        fileName = QFileDialog.getSaveFileName(directory=os.path.join(root,'Excel_data'))[0].split(".")[0] + ".xlsx"
        self.filePath.setText(fileName)

    def collect(self, caliGesture):
        self.disableBtns(True)
        fileName = self.filePath.text()
        self.thread = QThread()
        self.worker = CollectWorker(fileName, caliGesture)
        self.nextBtn.clicked.connect(lambda:self.worker.nextSignal.emit())
        self.backBtn.clicked.connect(lambda:self.worker.backSignal.emit())
        self.stopBtn.clicked.connect(lambda:self.worker.stopSignal.emit())
        self.worker.gestureSignal.connect(lambda gesture: self.gesture.setText(gesture))
        self.worker.resultSignal.connect(lambda values: self.values.setText(values))
        self.worker.finishSignal.connect(lambda: self.disableBtns(False))
        self.worker.finishSignal.connect(self.worker.deleteLater)
        self.worker.finishSignal.connect(self.thread.quit)
        self.thread.finished.connect(self.thread.deleteLater)
        self.worker.moveToThread(self.thread)
        self.thread.started.connect(self.worker.work)
        self.thread.start()

    def next(self):
        if self.nextBtn.isEnabled():
            self.worker.nextSignal.emit()

    def disableBtns(self, state):
        self.collectBtn.setDisabled(state)
        self.nextBtn.setDisabled(not state)
        self.backBtn.setDisabled(not state)
        self.stopBtn.setDisabled(not state)


class CollectWorker(QObject):
    gestureSignal = pyqtSignal(str)
    resultSignal = pyqtSignal(str)
    nextSignal = pyqtSignal()
    backSignal = pyqtSignal()
    stopSignal = pyqtSignal()
    finishSignal = pyqtSignal()

    def __init__(self, filePath, caliGesture):
        super().__init__()
        self.rf = Rf()
        self.next = False
        self.stop = False
        self.cnt = 0
        self.state = "Calibration"
        self.filePath = filePath
        self.caliGesture = caliGesture
        self.nextSignal.connect(lambda: setattr(self, "next", True))
        self.backSignal.connect(self.back)
        self.stopSignal.connect(lambda: setattr(self, "stop", True))

    def work(self):
        if not os.path.exists(self.filePath):
            workbook = Workbook()
            workbook.save(self.filePath)
            workbook.close()
        workbook = load_workbook(self.filePath)
        start = datetime.now()
        worksheet = workbook.create_sheet("calibration")
        title = ["gesture", "start", "end"]
        for i in range(len(title)):
            worksheet.cell(row=1, column=i + 1, value=title[i])
        for i in range(nSensor):
            worksheet.cell(row=1, column=20 * i + len(title) + 1, value=i)
        self.cnt = 0
        while self.cnt < caliCnt:
            self.gestureSignal.emit(f"Calibration {self.cnt+1} {gestures[self.caliGesture]}")
            while not self.next:
                pass
            worksheet.cell(row=self.cnt + 2, column=1, value=self.caliGesture)
            self.resultSignal.emit(
                self.rf.collect(worksheet, self.cnt + 2, start),
            )
            workbook.save(self.filePath)
            self.cnt += 1
            self.next = False
        start = datetime.now()
        worksheet = workbook.create_sheet("random")
        for i in range(len(title)):
            worksheet.cell(row=1, column=i + 1, value=title[i])
        for i in range(nSensor):
            worksheet.cell(row=1, column=20 * i + len(title) + 1, value=i)
        self.cnt = 0
        self.state = "Collect"
        while not self.stop:
            self.gestureSignal.emit(f"Collect {self.cnt+1} {gestures[self.cnt%len(gestures)]}")
            while not self.next:
                if self.stop:
                    break
            if self.stop:
                break
            worksheet.cell(row=self.cnt + 2, column=1, value=self.cnt % len(gestures))
            self.resultSignal.emit(
                self.rf.collect(worksheet, self.cnt + 2, start),
            )
            workbook.save(self.filePath)
            self.cnt += 1
            self.next = False
        workbook.close()
        self.finishSignal.emit()
    
    def back(self):
        self.cnt = max(self.cnt-1, 0)
        if self.state == "Calibration":
            self.gestureSignal.emit(f"Calibration {self.cnt+1} {gestures[self.caliGesture]}")
        else:
            self.gestureSignal.emit(f"Collect {self.cnt+1} {gestures[self.cnt%len(gestures)]}")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    font = QFont()
    font.setPointSize(13)
    app.setFont(font)
    MainWindow = Tool()
    sys.exit(app.exec_())
