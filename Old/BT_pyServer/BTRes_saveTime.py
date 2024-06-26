from pickle import NONE
import interface

from openpyxl import *
from datetime import datetime
from time import time

interf = interface.interface()

def main():
    fileName = 'D://NTU/Ben50/Excel_data/factor/base_QQ.xlsx'
    gestures = ['paper','rock']
    # gestures = ['nothing']
    workbook = load_workbook(fileName)
    title = ['gesture', 't', 'val']
    gesture=0
    for t in range(10):
        gesture=t%2
        print('ready '+gestures[gesture])
        readyTimer=time()
        worksheet = workbook.create_sheet(gestures[gesture])
        for i in range(len(title)):
            worksheet.cell(row=1, column=i+1, value=title[i])
        row = 2
        while(int((time()-readyTimer)*1000)<1000):
            pass
        print(str(t))
        interf.write(str(1))
        start = time()
        val = float(interf.read())
        while(val!=2048):
            worksheet.cell(row=row, column=1, value=gesture)
            worksheet.cell(row=row, column=2, value=int((time()-start)*1000))
            res=3000 * val / (5000 - val * 3)
            worksheet.cell(row=row, column=3, value=round(res, 2))
            row += 1
            print(res)
            val = float(interf.read())
        workbook.save(fileName)        
        print('relax')
        relaxTimer = time()
        while(int((time()-relaxTimer)*1000)<20000):
            pass
    interf.end_process()
    workbook.close()

if __name__ == '__main__':
    main()