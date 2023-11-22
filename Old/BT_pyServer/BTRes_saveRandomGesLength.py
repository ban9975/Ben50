from pickle import NONE
import interface
import sys

from openpyxl import *
from datetime import datetime

interf = interface.interface()
def res2len(x, z):
    p00 = -0.937767447539916
    p10 = -0.039948024365072
    p01 = 2.587084207187248
    p20 = 0.000614994711211
    p11 = 0.067406486859726
    p02 = -2.365525377675638
    p21 = -0.000536654708327
    p12 = -0.027106092579830
    p03 = 0.717608771791744

    err = sys.maxsize
    len = 1
    # print(z)
    for i in range(5000):
        y = 0.8 + 0.0001 * i
        tmp = (p00 + p10*x + p01*y + p20*(x**2) + p11*x*y + p02*(y**2) + p21*(x**2)*y + p12*x*(y**2) + p03*(y**3)) * 10**5
        tmpErr = abs(tmp - z)
        # print(tmp)
        if tmpErr < err:
            err = tmpErr
            len = y
    return len
def calibration(z):
    p00 = -0.937767447539916
    p10 = -0.039948024365072
    p01 = 2.587084207187248
    p20 = 0.000614994711211
    p11 = 0.067406486859726
    p02 = -2.365525377675638
    p21 = -0.000536654708327
    p12 = -0.027106092579830
    p03 = 0.717608771791744

    err = sys.maxsize
    len = 1.8
    y = 1
    # print(z)
    for i in range(30000):
        x = 1.8 + 0.0001 * i
        tmp = (p00 + p10*x + p01*y + p20*(x**2) + p11*x*y + p02*(y**2) + p21*(x**2)*y + p12*x*(y**2) + p03*(y**3)) * 10**5
        tmpErr = abs(tmp - z)
        # print(tmp)
        if tmpErr < err:
            err = tmpErr
            len = x
    return len
    
def main():
    fileName = 'wristband/small_adi_len_2.xlsx'
    gestures = ['down', 'up', 'thumb', 'little finger', 'stretch', 'fist', 'rest']
    workbook = load_workbook(fileName)
    start = datetime.now()
    mode = int(input("0: calibration, 1: random: "))
    stretch = [2.593161111111111, 2.2697611111111105, 2.653238888888889, 2.7099777777777785]
    print(stretch)
    if mode == 1:
        worksheet = workbook.create_sheet("random")
        title = ['gesture', 'start', 'end', 0, 1, 2, 3]
        for i in range(len(title)):
            worksheet.cell(row=1, column=i+1, value=title[i])
        for i in range(4):
            worksheet.cell(row=2, column=i+4, value=stretch[i])
    elif mode == 0:
        stretch = [0, 0, 0, 0]
    row = 3
    while True:
        gesture = (row-3)%7
        _input = input("{}\tinput: ".format(gestures[gesture]))
        if _input == 'b':
            row -= 1
        elif _input == 'e':
            interf.end_process()
            if mode == 1:
                workbook.save(fileName)
                workbook.close()
            elif mode == 0:
                for k in range(4):
                    stretch[k] = stretch[k]/(row - 3)
                print(stretch)
            break
        else:
            interf.write(str(1))
            if mode == 1:
                worksheet.cell(row=row, column=1, value=gesture)
                worksheet.cell(row=row, column=2, value=str(datetime.now()-start))

            avg = [0, 0, 0, 0]
            for j in range(20):
                for k in range(4):
                    btIn = float(interf.read())
                    while btIn == 0:
                        print(0)
                        btIn = float(interf.read())
                    avg[k] += btIn
            print(row-2, end='\t')
            for k in range(4):
                avg[k] /= 20
                res = 300 * avg[k] / (5000 - avg[k] * 3)
                
                if mode == 1:
                    leng = res2len(stretch[k],round(res,2))-1
                    print(leng, end='\t')                               
                    worksheet.cell(row=row, column=k+4, value=leng)
                    workbook.save(fileName)
                elif mode == 0:
                    leng = calibration(round(res,2))
                    print(leng, end='\t')                    
                    stretch[k] += leng

            print()
            if mode == 1:
                worksheet.cell(row=row, column=3, value=str(datetime.now()-start))
            row += 1

if __name__ == '__main__':
    main()